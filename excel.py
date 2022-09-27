from openpyxl import load_workbook


class excelRead:
    def readData():
        load_wb = load_workbook('/Users/seancho/Desktop/SandBox/Alphavantage/Book1.xlsx')

        load_ws = load_wb['Sheet1']
        # print(load_ws['A1'].value)

        # get_cells = load_ws['A1':'C2']
        # for row in get_cells:
        #     for cell in row:
        #         print(cell.value)

        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        # print(all_values)

        return all_values

    def readDataTradingView():
        load_wb = load_workbook('Book2.xlsx')

        load_ws = load_wb['Sheet1']
        # print(load_ws['A1'].value)

        # get_cells = load_ws['A1':'C2']
        # for row in get_cells:
        #     for cell in row:
        #         print(cell.value)

        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        # print(all_values)

        return all_values

    def readCurrentHoldings():
        load_wb = load_workbook('Book3.xlsx')

        load_ws = load_wb['Sheet1']
        # print(load_ws['A1'].value)

        # get_cells = load_ws['A1':'C2']
        # for row in get_cells:
        #     for cell in row:
        #         print(cell.value)

        all_values = []
        for row in load_ws.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        # print(all_values)

        return all_values
